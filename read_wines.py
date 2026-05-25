from pathlib import Path
import requests
from openpyxl import load_workbook

URL = 'https://dvmn.org/media/filer_public/5e/a4/5ea42013-ea14-4179-8c0f-25b0f7d2bcb1/wine.xlsx'
DATA_DIR = Path('data')
DATA_DIR.mkdir(exist_ok=True)
XLSX_PATH = DATA_DIR / 'wine.xlsx'


def download_xlsx(url: str, dest: Path) -> None:
    if dest.exists():
        print(f'Using existing file: {dest}')
        return
    print(f'Downloading {url} to {dest}...')
    resp = requests.get(url, stream=True)
    resp.raise_for_status()
    with open(dest, 'wb') as f:
        for chunk in resp.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
    print('Download complete.')


def read_wines(path: Path):
    print(f'Reading Excel file: {path}')
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() for h in next(rows)]
    except StopIteration:
        return []

    records = []
    for row in rows:
        if not any(cell is not None for cell in row):
            continue
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        records.append(rec)

    return records


if __name__ == '__main__':
    try:
        download_xlsx(URL, XLSX_PATH)
        wines = read_wines(XLSX_PATH)
        print(wines)
    except Exception as exc:
        print('Error:', exc)
        raise
