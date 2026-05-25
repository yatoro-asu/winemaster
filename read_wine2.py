from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List

import requests
from openpyxl import load_workbook

URL = (
    'https://dvmn.org/media/filer_public/ea/82/ea8233a3-a3ee-4b1e-ba72-50fec3f170ec/wine3.xlsx'
)
DATA_DIR = Path('data')
DATA_DIR.mkdir(exist_ok=True)
XLSX_PATH = DATA_DIR / 'wine3.xlsx'

CATEGORY_HEADERS = {'категория', 'category', 'categories'}


def download_xlsx(url: str, dest: Path) -> None:
    if dest.exists():
        print(f'Using existing file: {dest}')
        return
    print(f'Downloading {url} to {dest}...')
    resp = requests.get(url, stream=True)
    resp.raise_for_status()
    with open(dest, 'wb') as f:
        for chunk in resp.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
    print('Download complete.')


def normalize_value(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, float) and value != value:  # NaN
        return ''
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def read_excel_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else '' for cell in next(rows)]
    except StopIteration:
        return []

    records = []
    for row in rows:
        if not any(cell is not None for cell in row):
            continue
        record = {
            headers[i]: normalize_value(row[i])
            for i in range(min(len(headers), len(row)))
            if headers[i]
        }
        records.append(record)

    return records


def find_category_key(headers: Iterable[str]) -> str:
    lower_headers = {header.lower(): header for header in headers}
    for candidate in CATEGORY_HEADERS:
        if candidate in lower_headers:
            return lower_headers[candidate]
    raise ValueError('Category column not found in Excel headers.')


def group_by_category(records: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    if not records:
        return {}

    category_key = find_category_key(records[0].keys())
    grouped = defaultdict(list)
    for record in records:
        category = normalize_value(record.get(category_key, '')) or 'Без категории'
        grouped[category].append(record)
    return grouped


if __name__ == '__main__':
    try:
        download_xlsx(URL, XLSX_PATH)
        wines = read_excel_rows(XLSX_PATH)
        production = group_by_category(wines)
        print(production)
    except Exception as exc:
        print('Error:', exc)
        raise
